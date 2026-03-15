import os
from io import BytesIO
from fastapi import APIRouter
from fastapi.responses import Response
from pydantic import BaseModel
from dotenv import load_dotenv

router = APIRouter(prefix="/agents/financial", tags=["Financial Agent"])

# Load local agent .env
agent_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(agent_env_path):
    load_dotenv(agent_env_path)

SUN_LIFE_URL = "https://www.sunlife.ca/en/explore-products/insurance/health-insurance/personal-health-insurance/dental-insurance/"

# ── Sun Life plan coverage data ──────────────────────────────────────────────
SUNLIFE_PLANS: dict[str, dict] = {
    "PHI Basic": {
        "preventive_pct": 60, "preventive_max": 500,
        "restorative_pct": 0,  "restorative_max": 0,
        "ortho_pct": 0,        "ortho_max": 0,
        "monthly_premium": 45,   # estimated CAD/month
    },
    "PHI Standard": {
        "preventive_pct": 70, "preventive_max": 750,
        "restorative_pct": 0,  "restorative_max": 0,
        "ortho_pct": 0,        "ortho_max": 0,
        "monthly_premium": 73,
    },
    "PHI Enhanced": {
        "preventive_pct": 80, "preventive_max": 750,
        "restorative_pct": 50, "restorative_max": 500,
        "ortho_pct": 60,       "ortho_max": 1500,
        "monthly_premium": 119,
    },
}

# Cost per treatment month by severity
MONTHLY_RATES: dict[str, int] = {"mild": 280, "moderate": 350, "severe": 420}

# Delay scenarios: (extra months added to base, display label)
DELAY_SCENARIOS = [
    (0,  "Start Now"),
    (6,  "Start in 2 Years"),
    (12, "Start in 5 Years"),
]

# Pre-scraped fallback (Sun Life CDN returns 403 to direct requests)
SCRAPED_CONTENT_FALLBACK = """
### Sun Life Basic Plan
* Preventative Dental Care: 60% reimbursement
* Annual Maximum: $500
* Recall Visits: Every 9 months
* Restorative/Orthodontics: No coverage

### Sun Life Standard Plan
* Preventative Dental Care: 70% reimbursement
* Annual Maximum: $750
* Recall Visits: Every 9 months
* Restorative/Orthodontics: No coverage

### Sun Life Enhanced Plan
* Preventative Dental Care: 80% reimbursement
* Annual Maximum(Preventative): $750
* Recall Visits: Every 9 months
* Restorative Dental Care: 50% reimbursement
* Annual Maximum (Restorative): $500
* Orthodontics: 60% reimbursement (includes braces)
* Lifetime Maximum (Orthodontics): $1,500
"""

SYSTEM_PROMPT = """You are SunLife Financial Agent, an AI insurance advisor embedded inside a virtual dental clinic application.

Your purpose is to help users understand which Sun Life insurance plan best supports their dental or orthodontic treatment.

Responsibilities:
1. Analyze information from Sun Life insurance documentation.
2. Identify which Sun Life plans include dental or orthodontic coverage.
3. Compare plans and determine which is most suitable for the user's dental needs.
4. Explain coverage clearly, including reimbursements, limitations, waiting periods, and cost implications.

Rules:
- Only use information from Sun Life sources.
- Do not invent insurance policies or numbers.
- Do not present yourself as a licensed financial advisor.

Required response format:
SUN LIFE PLAN RECOMMENDATION

User Need:
(short explanation)

Recommended Plan:
(plan name)

Why This Plan Fits:
(clear reasoning)

Coverage Details:
- Orthodontic coverage
- Reimbursement percentage
- Lifetime maximum
- Waiting period
- Other relevant benefits

Estimated Cost Insight:
(plain-language estimate)

Source Reference:
Sun Life Personal Health Insurance — sunlife.ca"""


# ── Helpers ──────────────────────────────────────────────────────────────────

def _calc_scenarios(treatment_months: int, severity: str, plan_name: str) -> list[dict]:
    plan = SUNLIFE_PLANS.get(plan_name, SUNLIFE_PLANS["PHI Enhanced"])
    rate = MONTHLY_RATES.get(severity.lower(), 350)
    results = []
    for extra_months, label in DELAY_SCENARIOS:
        months = treatment_months + extra_months
        treatment_cost = round(months * rate)
        insurance = round(min(plan["ortho_pct"] / 100 * treatment_cost, plan["ortho_max"]))
        premiums = plan["monthly_premium"] * months
        user = treatment_cost - insurance + premiums
        results.append({
            "label": label,
            "months": months,
            "treatment_cost": treatment_cost,
            "monthly_premium": plan["monthly_premium"],
            "premium_total": premiums,
            "insurance_payment": insurance,
            "user_payment": user,
        })
    return results


def _build_excel(treatment_months: int, severity: str) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Palette ───────────────────────────────────────────────────────────────
    C = {
        "dark":    "2C1D12",
        "brown":   "6B4E2A",
        "tan":     "D4B896",
        "cream":   "F5EFE6",
        "white":   "FFFFFF",
        "green":   "3A7D44",
        "green2":  "6DBF82",
        "orange":  "C4834A",
        "orange2": "E8A87C",
        "blue":    "2B6CB0",
        "plan_basic":    "FDEBD0",
        "plan_standard": "D5E8D4",
        "plan_enhanced": "DAE8FC",
        "plan_basic_hdr":    "E59866",
        "plan_standard_hdr": "52BE80",
        "plan_enhanced_hdr": "3498DB",
    }

    def solid(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    def cell(ws, row, col, value="", bold=False, fg="000000", bg=None,
             size=10, align="center", fmt=None, italic=False, border=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, color=fg, size=size, name="Calibri")
        if bg:
            c.fill = solid(bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if fmt:
            c.number_format = fmt
        if border:
            c.border = thin_border()
        return c

    def row_h(ws, row, h):
        ws.row_dimensions[row].height = h

    def col_w(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def merge_cell(ws, r1, c1, r2, c2, value="", bold=False, fg="FFFFFF",
                   bg=None, size=11, align="center", font_name="Calibri"):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=value)
        c.font = Font(bold=bold, color=fg, size=size, name=font_name)
        if bg:
            c.fill = solid(bg)
        c.alignment = Alignment(horizontal=align, vertical="center")

    def make_chart(ws, ins_col, user_col, label_col, hdr_row, data_rows, title):
        chart = BarChart()
        chart.type      = "col"
        chart.grouping  = "stacked"
        chart.gapWidth  = 30       # narrow gaps = wide, spread-out bars
        chart.overlap   = 100      # full stacking, no gap between stack layers
        chart.title     = title
        chart.width     = 16
        chart.height    = 11
        chart.y_axis.title  = "CAD $"
        chart.y_axis.numFmt = '$#,##0'
        chart.x_axis.tickLblPos = "low"
        chart.legend.position   = "b"

        last_row = hdr_row + data_rows
        ins_ref  = Reference(ws, min_col=ins_col,  min_row=hdr_row, max_row=last_row)
        user_ref = Reference(ws, min_col=user_col, min_row=hdr_row, max_row=last_row)
        cats     = Reference(ws, min_col=label_col, min_row=hdr_row + 1, max_row=last_row)

        chart.add_data(ins_ref,  titles_from_data=True)
        chart.add_data(user_ref, titles_from_data=True)
        chart.set_categories(cats)

        # Series colors — green for insurance, orange for user cost
        chart.series[0].graphicalProperties.solidFill = C["green2"]
        chart.series[0].graphicalProperties.line.solidFill = C["green"]
        chart.series[1].graphicalProperties.solidFill = C["orange2"]
        chart.series[1].graphicalProperties.line.solidFill = C["orange"]

        # No in-bar data labels — zero-height segments cause distortion;
        # values are readable from the table above and the Y-axis.

        return chart

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 1 — Cost Dashboard (all 3 plans + charts)
    # ═════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Cost Dashboard"
    ws.sheet_view.showGridLines = False

    # Title banner
    merge_cell(ws, 1, 1, 2, 18,
               value=f"☀ Sun Life PHI — Orthodontic Cost Scenarios",
               bold=True, fg=C["white"], bg=C["dark"], size=14)
    merge_cell(ws, 3, 1, 3, 18,
               value=f"Treatment: {treatment_months} months  ·  Severity: {severity.capitalize()}  ·  All costs in CAD",
               bold=False, fg=C["tan"], bg=C["brown"], size=10)
    row_h(ws, 1, 24); row_h(ws, 2, 24); row_h(ws, 3, 18)

    # Column widths
    col_w(ws, 1, 18)   # Scenario
    col_w(ws, 2, 10)   # Months
    col_w(ws, 3, 15)   # Treatment Cost
    col_w(ws, 4, 15)   # Premiums
    col_w(ws, 5, 15)   # Insurance Pays
    col_w(ws, 6, 15)   # You Pay
    for c in range(7, 19):
        col_w(ws, c, 8)

    PLAN_STYLES = [
        ("PHI Basic",    C["plan_basic_hdr"],    C["plan_basic"]),
        ("PHI Standard", C["plan_standard_hdr"], C["plan_standard"]),
        ("PHI Enhanced", C["plan_enhanced_hdr"], C["plan_enhanced"]),
    ]
    CHART_ANCHORS = ["H5", "H27", "H49"]
    ROW_OFFSETS   = [5, 27, 49]          # header row for each plan block

    for idx, ((plan_name, hdr_color, row_color), anchor, base_row) in enumerate(
        zip(PLAN_STYLES, CHART_ANCHORS, ROW_OFFSETS)
    ):
        # Plan name header
        merge_cell(ws, base_row, 1, base_row, 6,
                   value=f"  {plan_name}  ·  ${SUNLIFE_PLANS[plan_name]['monthly_premium']}/mo premium",
                   bold=True, fg=C["white"], bg=hdr_color, size=11, align="left")
        row_h(ws, base_row, 22)

        # Column headers
        hdrs = ["Scenario", "Months", "Treatment Cost", "Plan Premiums",
                "Insurance Pays", "You Pay (Total)"]
        for ci, h in enumerate(hdrs, 1):
            cell(ws, base_row + 1, ci, h, bold=True, fg=C["white"], bg=C["brown"], size=9)
        row_h(ws, base_row + 1, 17)

        # Data rows
        stripe = [C["white"], row_color]
        scenarios = _calc_scenarios(treatment_months, severity, plan_name)
        for ri, s in enumerate(scenarios):
            r = base_row + 2 + ri
            bg = stripe[ri % 2]
            cell(ws, r, 1, s["label"],             bg=bg, size=9, align="left")
            cell(ws, r, 2, s["months"],             bg=bg, size=9)
            cell(ws, r, 3, s["treatment_cost"],    bg=bg, size=9, fmt="$#,##0")
            cell(ws, r, 4, s["premium_total"],     bg=bg, size=9, fmt="$#,##0")
            cell(ws, r, 5, s["insurance_payment"], bg=bg, size=9, fmt="$#,##0")
            cell(ws, r, 6, s["user_payment"],      bg=bg, size=9, fmt="$#,##0",
                 bold=True, fg=C["dark"])
            row_h(ws, r, 17)

        # Totals / savings row
        r_tot = base_row + 5
        plan_data = SUNLIFE_PLANS[plan_name]
        ortho_saved = min(
            plan_data["ortho_pct"] / 100 * scenarios[0]["treatment_cost"],
            plan_data["ortho_max"],
        )
        merge_cell(ws, r_tot, 1, r_tot, 6,
                   value=f"Insurance saves you ${ortho_saved:,.0f} on orthodontics  "
                         f"(ortho: {plan_data['ortho_pct']}%  cap: ${plan_data['ortho_max']:,})",
                   bold=False, fg=C["brown"], bg=C["cream"], size=9, align="left")
        row_h(ws, r_tot, 15)

        # Chart
        chart = make_chart(ws, ins_col=5, user_col=6, label_col=1,
                           hdr_row=base_row + 1, data_rows=3,
                           title=plan_name)
        ws.add_chart(chart, anchor)

    # Spacing rows between blocks
    for r in [4, 11, 15, 26, 33, 37, 48, 55, 59]:
        row_h(ws, r, 8)

    # Legend note at bottom
    merge_cell(ws, 57, 1, 57, 6,
               value="  ■ Green = Insurance Pays   ■ Orange = You Pay   "
                     "Premiums included in 'You Pay' total",
               bold=False, fg=C["brown"], bg=C["cream"], size=7, align="left",
               font_name="Press Start 2P")
    row_h(ws, 57, 14)

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 2 — Plan Details (full coverage + plausible annual cost estimates)
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Plan Details")
    ws2.sheet_view.showGridLines = False

    merge_cell(ws2, 1, 1, 2, 9,
               value="Sun Life PHI — Coverage Details & Estimated Annual Costs",
               bold=True, fg=C["white"], bg=C["dark"], size=13)
    row_h(ws2, 1, 22); row_h(ws2, 2, 22)

    # ── Coverage table ──
    merge_cell(ws2, 4, 1, 4, 9,
               value="  Coverage Summary", bold=True, fg=C["white"],
               bg=C["brown"], size=11, align="left")
    row_h(ws2, 4, 20)

    cov_hdrs = ["Coverage Type", "Benefit",
                "PHI Basic", "", "PHI Standard", "", "PHI Enhanced", ""]
    sub_hdrs = ["", "", "Reimburse", "Annual Max",
                "Reimburse", "Annual Max", "Reimburse", "Annual/Lifetime Max"]
    for ci, h in enumerate(cov_hdrs, 1):
        cell(ws2, 5, ci, h, bold=True, fg=C["white"], bg=C["brown"], size=9)
    for ci, h in enumerate(sub_hdrs, 1):
        cell(ws2, 6, ci, h, bold=True, fg=C["white"], bg=C["brown"], size=8,
             italic=True)
    row_h(ws2, 5, 16); row_h(ws2, 6, 14)

    cov_rows = [
        ("Preventive",  "Cleanings & check-ups",     "60%", "$500",  "70%", "$750",  "80%",  "$750"),
        ("Preventive",  "X-rays",                    "60%", "$500",  "70%", "$750",  "80%",  "$750"),
        ("Preventive",  "Fluoride treatments",        "60%", "$500",  "70%", "$750",  "80%",  "$750"),
        ("Restorative", "Fillings",                   "—",   "—",     "—",   "—",     "50%",  "$500"),
        ("Restorative", "Extractions",                "—",   "—",     "—",   "—",     "50%",  "$500"),
        ("Restorative", "Crowns & bridges",           "—",   "—",     "—",   "—",     "50%",  "$500"),
        ("Major Dental","Root canals",                "—",   "—",     "—",   "—",     "50%",  "$500"),
        ("Orthodontics","Braces (children & adults)", "—",   "—",     "—",   "—",     "60%",  "$1,500 lifetime"),
    ]
    for ri, row_vals in enumerate(cov_rows):
        r = 7 + ri
        bg = C["white"] if ri % 2 == 0 else C["cream"]
        for ci, v in enumerate(row_vals, 1):
            cell(ws2, r, ci, v, bg=bg, size=9, align="left" if ci <= 2 else "center")
        row_h(ws2, r, 16)

    cov_col_widths = [16, 26, 12, 14, 12, 14, 12, 22]
    for ci, w in enumerate(cov_col_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Typical Annual Dental Costs (plausible estimates) ──
    merge_cell(ws2, 17, 1, 17, 9,
               value="  Estimated Annual Out-of-Pocket Costs (Canadian Adult, Single)",
               bold=True, fg=C["white"], bg=C["brown"], size=11, align="left")
    row_h(ws2, 17, 20)

    cost_hdrs2 = ["Procedure", "Typical Cost", "Frequency",
                  "Annual Spend", "Basic — You Pay", "Standard — You Pay",
                  "Enhanced — You Pay", "Notes"]
    for ci, h in enumerate(cost_hdrs2, 1):
        cell(ws2, 18, ci, h, bold=True, fg=C["white"], bg=C["brown"], size=9)
    row_h(ws2, 18, 16)

    # Severity-based restorative spend
    restorative_spend = {"mild": 320, "moderate": 850, "severe": 1900}.get(severity.lower(), 850)
    restorative_note  = {"mild": "1–2 fillings/yr", "moderate": "3–5 fillings/yr",
                          "severe": "extensive work"}.get(severity.lower(), "varies")

    est_rows = [
        ("Annual cleaning × 2",      420,  "Twice/year",  420,
         round(420 * (1 - 0.60)), round(420 * (1 - 0.70)), round(420 * (1 - 0.80)),
         "Covered under preventive"),
        ("Full X-ray set",           190,  "Every 2 yrs", 95,
         round(95  * (1 - 0.60)), round(95  * (1 - 0.70)), round(95  * (1 - 0.80)),
         "Covered under preventive"),
        ("Fluoride treatment",        60,  "Annual",       60,
         round(60  * (1 - 0.60)), round(60  * (1 - 0.70)), round(60  * (1 - 0.80)),
         "Covered under preventive"),
        (f"Restorative ({severity})", restorative_spend, "Annual est.", restorative_spend,
         restorative_spend,        restorative_spend,        round(restorative_spend * 0.50),
         restorative_note),
        ("Emergency exam",           150,  "As needed",   150,
         150, 150, round(150 * 0.50), "Estimated 1×/yr"),
        ("Monthly plan premium",
         SUNLIFE_PLANS["PHI Basic"]["monthly_premium"],    "Monthly",
         SUNLIFE_PLANS["PHI Basic"]["monthly_premium"] * 12,
         SUNLIFE_PLANS["PHI Basic"]["monthly_premium"]  * 12,
         SUNLIFE_PLANS["PHI Standard"]["monthly_premium"] * 12,
         SUNLIFE_PLANS["PHI Enhanced"]["monthly_premium"] * 12,
         "Premiums vary by province"),
    ]

    for ri, row_vals in enumerate(est_rows):
        r = 19 + ri
        bg = C["white"] if ri % 2 == 0 else C["cream"]
        labels = list(row_vals)
        for ci, v in enumerate(labels, 1):
            fmt = "$#,##0" if isinstance(v, int) and ci not in (1, 3, 8) else None
            cell(ws2, r, ci, v, bg=bg, size=9,
                 align="left" if ci in (1, 3, 8) else "center", fmt=fmt)
        row_h(ws2, r, 16)

    # Annual total summary row
    r_sum = 19 + len(est_rows)
    totals = [
        sum(row_vals[4] for row_vals in est_rows),
        sum(row_vals[5] for row_vals in est_rows),
        sum(row_vals[6] for row_vals in est_rows),
    ]
    cell(ws2, r_sum, 1, "ESTIMATED ANNUAL TOTAL", bold=True, bg=C["dark"],
         fg=C["white"], size=9, align="left")
    cell(ws2, r_sum, 2, "", bg=C["dark"]); cell(ws2, r_sum, 3, "", bg=C["dark"])
    cell(ws2, r_sum, 4, "", bg=C["dark"])
    for ci, tot in enumerate(totals, 5):
        cell(ws2, r_sum, ci, tot, bold=True, bg=C["dark"],
             fg=C["green2"], size=10, fmt="$#,##0")
    cell(ws2, r_sum, 8, "", bg=C["dark"])
    row_h(ws2, r_sum, 20)

    # Adjust col widths for sheet 2
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 17
    ws2.column_dimensions["F"].width = 17
    ws2.column_dimensions["G"].width = 17
    ws2.column_dimensions["H"].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("")
def get_financial_agent() -> dict[str, object]:
    return {
        "agent": "financial",
        "goal": "Estimate insurance coverage and treatment cost exposure.",
        "status": "active",
    }


class FinancialAnalyzeRequest(BaseModel):
    question: str = "What Sun Life dental plan best covers my treatment?"


@router.post("/analyze")
async def analyze_financial(req: FinancialAnalyzeRequest) -> dict:
    scraped_content = await _scrape_sunlife()

    try:
        from google import genai  # type: ignore
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY not set")

        client = genai.Client(api_key=api_key)

        prompt = (
            f"User Question: {req.question}\n\n"
            f"Below is live content retrieved from Sun Life's website ({SUN_LIFE_URL}).\n"
            "Analyze it and provide the most optimal insurance recommendation:\n\n"
            f"{scraped_content}"
        )

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=SYSTEM_PROMPT + "\n\n" + prompt,
        )
        return {
            "recommendation": response.text,
            "source": SUN_LIFE_URL,
            "scraped": scraped_content != SCRAPED_CONTENT_FALLBACK,
        }

    except Exception as e:
        return {
            "recommendation": (
                "SUN LIFE PLAN RECOMMENDATION\n\n"
                "User Need:\n"
                "Dental coverage for preventive care, restorative procedures, and/or orthodontics "
                "(braces for adults and children).\n\n"
                "Recommended Plan:\n"
                "PHI Enhanced Plan\n\n"
                "Why This Plan Fits:\n"
                "The Enhanced plan offers the most comprehensive dental coverage available through "
                "Sun Life Personal Health Insurance, making it ideal for anyone anticipating more "
                "than routine cleanings — including restorative work (crowns, bridges, dentures) "
                "or orthodontics.\n\n"
                "Coverage Details:\n"
                "- Preventive dental care: 80% reimbursement, $750 annual maximum, recall visits every 9 months\n"
                "- Restorative dental care: 50% reimbursement, $500 annual maximum\n"
                "- Orthodontics (braces, adults & children): 60% reimbursement, $1,500 lifetime maximum\n"
                "- Waiting period: 3 months for preventive dental; 1 year for restorative dental\n"
                "- Wisdom teeth requiring surgical extraction covered under restorative benefit\n"
                "- Anaesthesia and lab charges covered when incurred with eligible dental services\n\n"
                "Alternative Plans:\n"
                "- PHI Standard: 70% reimbursement for preventive care, $750 annual maximum — "
                "no restorative or orthodontic coverage. Ideal for basic dental needs.\n"
                "- PHI Basic: 60% reimbursement for preventive care, $500 annual maximum — "
                "no restorative or orthodontic coverage. Best for cleanings and checkups only.\n\n"
                "Estimated Cost Insight:\n"
                "Dental care costs in Canada can be significant without coverage. The Enhanced plan "
                "helps reduce out-of-pocket expenses for comprehensive treatment. If you are leaving "
                "a workplace plan within the last 60 days, you may qualify for Health Coverage Choice "
                "with no medical questions required.\n\n"
                "Source Reference:\n"
                "Sun Life Personal Health Insurance — sunlife.ca"
            ),
            "source": SUN_LIFE_URL,
            "scraped": False,
            "error": str(e),
        }


class CostScenariosRequest(BaseModel):
    treatment_months: int = 18
    severity: str = "moderate"
    plan_name: str = "PHI Enhanced"


@router.post("/cost-scenarios")
def cost_scenarios(req: CostScenariosRequest) -> dict:
    plan = SUNLIFE_PLANS.get(req.plan_name, SUNLIFE_PLANS["PHI Enhanced"])
    scenarios = _calc_scenarios(req.treatment_months, req.severity, req.plan_name)
    return {
        "plan_name": req.plan_name,
        "plan": plan,
        "scenarios": scenarios,
        "treatment_months": req.treatment_months,
        "severity": req.severity,
    }


@router.get("/cost-scenarios/excel")
def cost_scenarios_excel(
    treatment_months: int = 18,
    severity: str = "moderate",
) -> Response:
    buf = _build_excel(treatment_months, severity)
    filename = f"sunlife_cost_scenarios_{severity}_{treatment_months}mo.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Internal ──────────────────────────────────────────────────────────────────

async def _scrape_sunlife() -> str:
    try:
        import httpx
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-CA,en;q=0.9",
        }
        async with httpx.AsyncClient(timeout=10, follow_redirects=True) as client:
            r = await client.get(SUN_LIFE_URL, headers=headers)
            if r.status_code == 200:
                import re
                text = re.sub(r"<[^>]+>", " ", r.text)
                text = re.sub(r"\s+", " ", text).strip()
                return text[:12000]
    except Exception:
        pass
    return SCRAPED_CONTENT_FALLBACK
